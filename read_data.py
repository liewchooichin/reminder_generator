# -*- coding: utf-8 -*-
"""
Created on Tue Jan 17 21:50:07 2023

@author: Kang Liew Bei
"""


from openpyxl import load_workbook

class DataFile:

    def __init__(self, data_filename = '', template_filename = '', year_level=''):
        self.data_filename = data_filename
        self.template_filename = template_filename
        self.year_level = year_level
        self.record_dict = dict()

        try:
            self.data_wb = load_workbook(self.data_filename, read_only = True, data_only = True)
            self.data_ws = self.data_wb.active

            self.template_wb = load_workbook(self.template_filename)
            self.template_ws = self.template_wb.active
            print(f'template name: {self.template_ws.title}')
        except FileNotFoundError:
            print('File not found: ')


    def read_row_per_student(self):

        # print i num of rows
        new_name = ''
        keep_track_name = ''
        home_group = ''
        data = [['title1', 'due_date1']]
        first_record_flag = True

        # for value in row:
        """ col 1: home group """
        """ col 2: student name """
        """ col 3: title """
        """ col 4: due date """

        """ self.data_ws.values """
        """ is object of type 'generator'. """
        """ It has no len() """

        for row in self.data_ws.values:

            home_group = row[0]
            new_name = row[1]
            title = row[2]
            due_date = row[3]

            """ for every new name, a home group is appended. """
            if new_name not in self.record_dict:
                self.record_dict[new_name] = [ [home_group] ]
            """ if a name is already created (name, home_group), """
            """ then, append the (title, due_date). """
            if new_name in self.record_dict:
                self.record_dict[new_name].append([title, due_date])
        """ end for loop """

        #print(self.record_dict)
        self.write_rows()
        """ end def function """

    def write_rows(self):
        """ Name at C7 """
        """ Class at C9 """
        """ Title at B12 """
        """ Due Date at F12 """
        sheet_name = ''
        NAME_CELL = 'C7'
        HOME_GROUP_CELL = 'C9'
        TITLE_COL = 'B'
        DUE_DATE_COL = 'F'
        START_ROW_NUM = 12 # the starting row number in xlsx

        for k, v in self.record_dict.items():
            student_name = k
            home_group = v[0][0]
            sheet_name = home_group[0:3] + '_' + student_name[0:6]
            new_ws = self.template_wb.copy_worksheet(self.template_ws)
            new_ws.title = sheet_name
            new_ws[NAME_CELL] = student_name
            new_ws[HOME_GROUP_CELL] = self.year_level + ' ' + home_group
            for r in range(1, len(v)):
                ROW_NUM = START_ROW_NUM + r
                TITLE_CELL = TITLE_COL + str(ROW_NUM)
                DUE_DATE_CELL = DUE_DATE_COL + str(ROW_NUM)
                new_ws[TITLE_CELL] = v[r][0]
                new_ws[DUE_DATE_CELL] = v[r][1]


    def save_file(self):
        save_as_filename = self.year_level + '_reminder.xlsx'
        self.template_wb.save(save_as_filename)